"""
报告生成服务
Contract Review AI - Report Generator

生成多种格式的评审报告：
1. Markdown 管理层摘要报告
2. Excel 综合评审报告
3. CSV 风险矩阵
4. HTML 专业网页报告 (新增)
5. ZIP 打包下载
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import csv
import zipfile

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    评审报告生成器
    
    将AI评审结果转换为多种格式的专业报告
    """
    
    def __init__(self):
        self.risk_level_map = {
            "高": {"color": "🔴", "priority": 1},
            "中": {"color": "🟡", "priority": 2},
            "低": {"color": "🟢", "priority": 3}
        }
        
        self.status_map = {
            "通过": "✅",
            "不通过": "❌",
            "需关注": "⚠️"
        }
        
        # 初始化 HTML 报告生成器 (使用鹰眼风格 v2)
        try:
            from app.services.html_report_generator_v2 import HTMLReportGeneratorV2
            self.html_generator = HTMLReportGeneratorV2()
            logger.info("已加载鹰眼风格 HTML 报告生成器 v2")
        except ImportError:
            # 降级到 v1 版本
            try:
                from app.services.html_report_generator import HTMLReportGenerator
                self.html_generator = HTMLReportGenerator()
                logger.info("已加载 HTML 报告生成器 v1")
            except ImportError:
                self.html_generator = None
                logger.warning("HTMLReportGenerator 未找到，跳过HTML报告生成")
    
    def generate_all_reports(
        self,
        task_id: str,
        review_result: Dict[str, Any],
        output_dir: Path,
        contract_name: str = "合同"
    ) -> Dict[str, Path]:
        """
        生成所有格式的报告
        
        Args:
            task_id: 任务ID
            review_result: AI评审结果字典
            output_dir: 输出目录
            contract_name: 合同名称
            
        Returns:
            生成的报告文件路径字典
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        generated_files = {}
        
        try:
            # 1. 生成Markdown摘要报告
            md_path = self.generate_markdown_summary(
                review_result, output_dir, contract_name
            )
            generated_files["markdown_summary"] = md_path
            logger.info(f"已生成Markdown报告: {md_path}")
            
            # 2. 生成Excel综合报告
            excel_path = self.generate_excel_report(
                review_result, output_dir, contract_name
            )
            generated_files["excel_report"] = excel_path
            logger.info(f"已生成Excel报告: {excel_path}")
            
            # 3. 生成风险矩阵CSV
            risk_csv_path = self.generate_risk_matrix_csv(
                review_result, output_dir
            )
            generated_files["risk_matrix"] = risk_csv_path
            logger.info(f"已生成风险矩阵: {risk_csv_path}")
            
            # 4. 生成合规检查CSV
            compliance_csv_path = self.generate_compliance_csv(
                review_result, output_dir
            )
            generated_files["compliance_check"] = compliance_csv_path
            logger.info(f"已生成合规检查: {compliance_csv_path}")
            
            # 5. 生成HTML专业报告 (新增)
            if self.html_generator:
                html_path = self.html_generator.generate_html_report(
                    review_result=review_result,
                    output_path=output_dir / "review_report.html",
                    contract_name=contract_name,
                    task_id=task_id
                )
                generated_files["html_report"] = html_path
                logger.info(f"已生成HTML报告: {html_path}")
            
            # 6. 打包所有报告为ZIP
            zip_path = self.create_zip_package(
                task_id, generated_files, output_dir
            )
            generated_files["zip_package"] = zip_path
            logger.info(f"已生成ZIP包: {zip_path}")
            
        except Exception as e:
            logger.error(f"生成报告时出错: {e}", exc_info=True)
            raise
        
        return generated_files
    
    def generate_markdown_summary(
        self,
        review_result: Dict[str, Any],
        output_dir: Path,
        contract_name: str = "合同"
    ) -> Path:
        """
        生成Markdown格式的管理层摘要报告
        """
        output_path = output_dir / "management_summary.md"
        
        # 提取数据
        overall = review_result.get("overall_assessment", "")
        risk_level = review_result.get("risk_level", "未知")
        key_findings = review_result.get("key_findings", [])
        compliance = review_result.get("compliance_check", [])
        recommendations = review_result.get("recommendations", [])
        missing_clauses = review_result.get("missing_clauses", [])
        
        # 计算统计数据
        risk_icon = self.risk_level_map.get(risk_level, {}).get("color", "⚪")
        high_risk_count = sum(1 for f in key_findings if f.get("severity") == "高")
        medium_risk_count = sum(1 for f in key_findings if f.get("severity") == "中")
        low_risk_count = sum(1 for f in key_findings if f.get("severity") == "低")
        
        compliance_pass = sum(1 for c in compliance if c.get("status") == "通过")
        compliance_fail = sum(1 for c in compliance if c.get("status") == "不通过")
        compliance_warn = sum(1 for c in compliance if c.get("status") == "需关注")
        
        # 生成报告内容
        content = f"""# 合同评审报告 - 管理层摘要

**报告生成时间**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  
**合同名称**: {contract_name}  
**评审状态**: 已完成

---

## 一、评审结论

### 整体风险等级: {risk_icon} {risk_level}

{overall}

### 风险分布统计

| 风险等级 | 数量 |
|---------|------|
| 🔴 高风险 | {high_risk_count} 项 |
| 🟡 中风险 | {medium_risk_count} 项 |
| 🟢 低风险 | {low_risk_count} 项 |
| **合计** | **{len(key_findings)} 项** |

### 合规检查统计

| 检查结果 | 数量 |
|---------|------|
| ✅ 通过 | {compliance_pass} 项 |
| ❌ 不通过 | {compliance_fail} 项 |
| ⚠️ 需关注 | {compliance_warn} 项 |
| **合计** | **{len(compliance)} 项** |

---

## 二、关键风险发现

"""
        
        # 添加关键风险发现
        for i, finding in enumerate(key_findings, 1):
            severity = finding.get("severity", "未知")
            severity_icon = self.risk_level_map.get(severity, {}).get("color", "⚪")
            
            content += f"""### {i}. {finding.get("category", "未分类")} {severity_icon}

**风险等级**: {severity}  
**问题位置**: {finding.get("location", "未指定")}

**问题描述**:  
{finding.get("description", "")}

**改进建议**:  
{finding.get("suggestion", "")}

---

"""
        
        # 添加合规检查结果
        content += """## 三、合规检查结果

| 检查项目 | 状态 | 说明 |
|---------|------|------|
"""
        
        for check in compliance:
            status = check.get("status", "未知")
            status_icon = self.status_map.get(status, "❓")
            content += f"| {check.get('item', '')} | {status_icon} {status} | {check.get('details', '')} |\n"
        
        # 添加修改建议
        content += """
---

## 四、修改建议

"""
        for i, rec in enumerate(recommendations, 1):
            content += f"{i}. {rec}\n\n"
        
        # 添加缺失条款
        if missing_clauses:
            content += """---

## 五、缺失条款清单

以下条款在合同中缺失，建议补充：

"""
            for i, clause in enumerate(missing_clauses, 1):
                content += f"{i}. {clause}\n"
        
        # 添加声明
        content += """
---

## 附录：免责声明

本报告由AI智能评审系统自动生成，仅供参考。建议在签署合同前，由专业法律人员进行最终审核。

---

*报告由 Contract Review AI 系统生成*
"""
        
        # 写入文件
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        
        return output_path
    
    def generate_excel_report(
        self,
        review_result: Dict[str, Any],
        output_dir: Path,
        contract_name: str = "合同"
    ) -> Path:
        """
        生成Excel格式的综合评审报告
        """
        output_path = output_dir / "comprehensive_report.xlsx"
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl未安装，跳过Excel报告生成")
            # 创建一个简单的CSV作为替代
            return self._generate_excel_as_csv(review_result, output_dir, contract_name)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 定义样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        
        risk_fills = {
            "高": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "中": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
            "低": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
        }
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Sheet 1: 摘要 ===
        ws_summary = wb.active
        ws_summary.title = "评审摘要"
        
        # 标题
        ws_summary['A1'] = "合同评审报告"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "报告生成时间"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary['A4'] = "合同名称"
        ws_summary['B4'] = contract_name
        ws_summary['A5'] = "整体风险等级"
        ws_summary['B5'] = review_result.get("risk_level", "未知")
        
        # 整体评估
        ws_summary['A7'] = "整体评估"
        ws_summary['A7'].font = header_font
        ws_summary['A8'] = review_result.get("overall_assessment", "")
        ws_summary.merge_cells('A8:D8')
        ws_summary['A8'].alignment = Alignment(wrap_text=True)
        ws_summary.row_dimensions[8].height = 80
        
        # 设置列宽
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50
        ws_summary.column_dimensions['C'].width = 20
        ws_summary.column_dimensions['D'].width = 30
        
        # === Sheet 2: 风险发现 ===
        ws_risks = wb.create_sheet("风险发现")
        
        # 表头
        headers = ["序号", "风险类别", "风险等级", "问题位置", "问题描述", "改进建议"]
        for col, header in enumerate(headers, 1):
            cell = ws_risks.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        for row, finding in enumerate(review_result.get("key_findings", []), 2):
            ws_risks.cell(row=row, column=1, value=row-1).border = border
            ws_risks.cell(row=row, column=2, value=finding.get("category", "")).border = border
            
            severity_cell = ws_risks.cell(row=row, column=3, value=finding.get("severity", ""))
            severity_cell.border = border
            severity_cell.fill = risk_fills.get(finding.get("severity", ""), PatternFill())
            
            ws_risks.cell(row=row, column=4, value=finding.get("location", "")).border = border
            
            desc_cell = ws_risks.cell(row=row, column=5, value=finding.get("description", ""))
            desc_cell.border = border
            desc_cell.alignment = Alignment(wrap_text=True)
            
            sugg_cell = ws_risks.cell(row=row, column=6, value=finding.get("suggestion", ""))
            sugg_cell.border = border
            sugg_cell.alignment = Alignment(wrap_text=True)
            
            ws_risks.row_dimensions[row].height = 60
        
        # 设置列宽
        widths = [8, 15, 12, 25, 50, 50]
        for i, width in enumerate(widths, 1):
            ws_risks.column_dimensions[get_column_letter(i)].width = width
        
        # === Sheet 3: 合规检查 ===
        ws_compliance = wb.create_sheet("合规检查")
        
        headers = ["序号", "检查项目", "检查状态", "详细说明"]
        for col, header in enumerate(headers, 1):
            cell = ws_compliance.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
        
        for row, check in enumerate(review_result.get("compliance_check", []), 2):
            ws_compliance.cell(row=row, column=1, value=row-1).border = border
            ws_compliance.cell(row=row, column=2, value=check.get("item", "")).border = border
            
            status_cell = ws_compliance.cell(row=row, column=3, value=check.get("status", ""))
            status_cell.border = border
            if check.get("status") == "通过":
                status_cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
            elif check.get("status") == "不通过":
                status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif check.get("status") == "需关注":
                status_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            
            detail_cell = ws_compliance.cell(row=row, column=4, value=check.get("details", ""))
            detail_cell.border = border
            detail_cell.alignment = Alignment(wrap_text=True)
            
            ws_compliance.row_dimensions[row].height = 40
        
        widths = [8, 30, 15, 60]
        for i, width in enumerate(widths, 1):
            ws_compliance.column_dimensions[get_column_letter(i)].width = width
        
        # === Sheet 4: 修改建议 ===
        ws_recommendations = wb.create_sheet("修改建议")
        
        headers = ["序号", "建议内容"]
        for col, header in enumerate(headers, 1):
            cell = ws_recommendations.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
        
        for row, rec in enumerate(review_result.get("recommendations", []), 2):
            ws_recommendations.cell(row=row, column=1, value=row-1).border = border
            rec_cell = ws_recommendations.cell(row=row, column=2, value=rec)
            rec_cell.border = border
            rec_cell.alignment = Alignment(wrap_text=True)
            ws_recommendations.row_dimensions[row].height = 50
        
        ws_recommendations.column_dimensions['A'].width = 8
        ws_recommendations.column_dimensions['B'].width = 100
        
        # === Sheet 5: 缺失条款 ===
        ws_missing = wb.create_sheet("缺失条款")
        
        headers = ["序号", "缺失条款"]
        for col, header in enumerate(headers, 1):
            cell = ws_missing.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
        
        for row, clause in enumerate(review_result.get("missing_clauses", []), 2):
            ws_missing.cell(row=row, column=1, value=row-1).border = border
            ws_missing.cell(row=row, column=2, value=clause).border = border
        
        ws_missing.column_dimensions['A'].width = 8
        ws_missing.column_dimensions['B'].width = 60
        
        # 保存工作簿
        wb.save(output_path)
        
        return output_path
    
    def _generate_excel_as_csv(
        self,
        review_result: Dict[str, Any],
        output_dir: Path,
        contract_name: str
    ) -> Path:
        """
        如果openpyxl不可用，生成CSV作为替代
        """
        output_path = output_dir / "comprehensive_report.csv"
        
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            
            # 摘要信息
            writer.writerow(["合同评审报告"])
            writer.writerow(["报告时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            writer.writerow(["合同名称", contract_name])
            writer.writerow(["风险等级", review_result.get("risk_level", "")])
            writer.writerow(["整体评估", review_result.get("overall_assessment", "")])
            writer.writerow([])
            
            # 风险发现
            writer.writerow(["风险发现"])
            writer.writerow(["序号", "类别", "等级", "位置", "描述", "建议"])
            for i, f in enumerate(review_result.get("key_findings", []), 1):
                writer.writerow([
                    i,
                    f.get("category", ""),
                    f.get("severity", ""),
                    f.get("location", ""),
                    f.get("description", ""),
                    f.get("suggestion", "")
                ])
            writer.writerow([])
            
            # 合规检查
            writer.writerow(["合规检查"])
            writer.writerow(["序号", "检查项", "状态", "说明"])
            for i, c in enumerate(review_result.get("compliance_check", []), 1):
                writer.writerow([
                    i,
                    c.get("item", ""),
                    c.get("status", ""),
                    c.get("details", "")
                ])
        
        return output_path
    
    def generate_risk_matrix_csv(
        self,
        review_result: Dict[str, Any],
        output_dir: Path
    ) -> Path:
        """
        生成风险矩阵CSV文件
        """
        output_path = output_dir / "risk_matrix.csv"
        
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            
            # 表头
            writer.writerow([
                "序号", "风险类别", "风险等级", "问题位置", 
                "问题描述", "改进建议", "优先级"
            ])
            
            # 按风险等级排序
            findings = review_result.get("key_findings", [])
            sorted_findings = sorted(
                findings,
                key=lambda x: self.risk_level_map.get(x.get("severity", ""), {}).get("priority", 99)
            )
            
            for i, finding in enumerate(sorted_findings, 1):
                severity = finding.get("severity", "未知")
                priority = self.risk_level_map.get(severity, {}).get("priority", 99)
                
                writer.writerow([
                    i,
                    finding.get("category", ""),
                    severity,
                    finding.get("location", ""),
                    finding.get("description", ""),
                    finding.get("suggestion", ""),
                    f"P{priority}"
                ])
        
        return output_path
    
    def generate_compliance_csv(
        self,
        review_result: Dict[str, Any],
        output_dir: Path
    ) -> Path:
        """
        生成合规检查CSV文件
        """
        output_path = output_dir / "compliance_check.csv"
        
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            
            # 表头
            writer.writerow(["序号", "检查项目", "检查状态", "详细说明"])
            
            for i, check in enumerate(review_result.get("compliance_check", []), 1):
                writer.writerow([
                    i,
                    check.get("item", ""),
                    check.get("status", ""),
                    check.get("details", "")
                ])
        
        return output_path
    
    def create_zip_package(
        self,
        task_id: str,
        generated_files: Dict[str, Path],
        output_dir: Path
    ) -> Path:
        """
        将所有报告打包为ZIP文件
        """
        zip_path = output_dir / f"review_reports_{task_id[:8]}.zip"
        
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, file_path in generated_files.items():
                if file_path and file_path.exists() and name != "zip_package":
                    zf.write(file_path, file_path.name)
            
            # 也包含原始JSON结果
            json_path = output_dir / "review_result.json"
            if json_path.exists():
                zf.write(json_path, json_path.name)
        
        return zip_path


# 单元测试
if __name__ == "__main__":
    # 测试数据
    test_result = {
        "overall_assessment": "该技术服务合同框架基本完整，但存在多项风险需要关注。",
        "risk_level": "高",
        "key_findings": [
            {
                "category": "交付风险",
                "severity": "高",
                "description": "交付分工界面不明确",
                "location": "第二条 服务内容",
                "suggestion": "明确交付物清单"
            },
            {
                "category": "知识产权风险",
                "severity": "中",
                "description": "未约定知识产权归属",
                "location": "全文缺失",
                "suggestion": "增加知识产权条款"
            }
        ],
        "compliance_check": [
            {"item": "合同主体信息", "status": "通过", "details": "信息完整"},
            {"item": "知识产权条款", "status": "不通过", "details": "完全缺失"}
        ],
        "recommendations": [
            "建议1：补充知识产权条款",
            "建议2：明确交付物清单"
        ],
        "missing_clauses": ["知识产权条款", "保密条款"]
    }
    
    # 生成报告
    generator = ReportGenerator()
    output_dir = Path("./test_reports")
    
    files = generator.generate_all_reports(
        task_id="test-001",
        review_result=test_result,
        output_dir=output_dir,
        contract_name="测试合同"
    )
    
    print("生成的报告文件:")
    for name, path in files.items():
        print(f"  {name}: {path}")
