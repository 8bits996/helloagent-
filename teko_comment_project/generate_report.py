# -*- coding: utf-8 -*-
"""
AI Agent大赛综合报告生成器
整合所有数据源：清单、评分、评论报告等
"""

import csv
import re
import json
from datetime import datetime

# ============ 1. 读取部门赛清单数据 ============
def read_dept_list():
    """读取部门赛清单"""
    data = {}
    with open('dept_agent_list.csv', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            agent_id = row.get('id', '').strip()
            if agent_id:
                data[agent_id] = {
                    'id': agent_id,
                    'agent名': row.get('agent名', '').strip(),
                    '体验链接': row.get('体验链接', '').strip(),
                    '主导部门': row.get('主导部门', '').strip(),
                    '涉及部门': row.get('涉及部门', '').strip(),
                    '项目经理': row.get('项目经理', '').strip(),
                    '项目成员': row.get('项目成员', '').strip(),
                }
    return data

# ============ 2. 读取评分数据 ============
def read_experience_data():
    """读取体验评分数据"""
    data = {}
    try:
        with open('experience_data.csv', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get('项目名称', '').strip()
                if name:
                    # 提取teko链接中的ID
                    teko_link = row.get('项目材料', '')
                    agent_id = ''
                    if 'ai-agent/' in teko_link:
                        match = re.search(r'ai-agent/(\d+)', teko_link)
                        if match:
                            agent_id = match.group(1)
                    
                    data[name] = {
                        'teko链接': teko_link,
                        'agent_id': agent_id,
                        '综合评分': row.get('*综合打分（0-100分）\n打分建议不要出现相同分数', '').strip(),
                        '评委建议': row.get('评委建议', '').strip(),
                        '特点': row.get('特点', '').strip(),
                        '访问链接': row.get('访问链接', '').strip(),
                        '技术方案': row.get('技术方案', '').strip(),
                    }
    except Exception as e:
        print(f"读取评分数据出错: {e}")
    return data

# ============ 3. 解析评论报告中的详细信息 ============
def parse_comment_report():
    """解析teko_comment_report.md中的详细信息"""
    data = {}
    try:
        with open('teko_comment_report.md', encoding='utf-8') as f:
            content = f.read()
        
        # 按作品块分割
        blocks = re.split(r'### \d+\. ', content)
        
        for block in blocks[1:]:  # 跳过第一个空块
            lines = block.strip().split('\n')
            if not lines:
                continue
            
            # 获取作品名
            name = lines[0].strip()
            if not name:
                continue
            
            info = {'作品名': name}
            
            # 解析表格中的字段
            for line in lines:
                if '|' in line and '**' in line:
                    # 提取字段名和值
                    match = re.search(r'\*\*(.+?)\*\*\s*\|\s*(.+)', line)
                    if match:
                        field = match.group(1).strip()
                        value = match.group(2).strip()
                        if value and value != '-':
                            info[field] = value
            
            # 用作品ID作为key
            if '作品ID' in info:
                data[info['作品ID']] = info
            else:
                # 尝试从参赛链接提取ID
                if '参赛链接' in info:
                    match = re.search(r'ai-agent/(\d+)', info['参赛链接'])
                    if match:
                        data[match.group(1)] = info
    except Exception as e:
        print(f"解析评论报告出错: {e}")
    return data

# ============ 4. 读取缺少信息的作品 ============
def read_missing_info():
    """读取缺少信息的作品"""
    data = {}
    try:
        with open('missing_info_agents.csv', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                agent_id = row.get('id', '').strip()
                if agent_id:
                    data[agent_id] = row.get('缺少字段', '').strip()
    except:
        pass
    return data

# ============ 5. 整合所有数据 ============
def merge_all_data():
    """整合所有数据源"""
    # 读取各数据源
    dept_list = read_dept_list()
    experience_data = read_experience_data()
    comment_data = parse_comment_report()
    missing_info = read_missing_info()
    
    # 建立评分数据的ID映射
    score_by_id = {}
    for name, info in experience_data.items():
        if info.get('agent_id'):
            score_by_id[info['agent_id']] = info
            score_by_id[info['agent_id']]['项目名称'] = name
    
    # 整合数据
    merged = {}
    for agent_id, info in dept_list.items():
        merged[agent_id] = {
            'id': agent_id,
            'agent名': info['agent名'],
            '体验链接': info['体验链接'],
            '主导部门': info['主导部门'],
            '涉及部门': info['涉及部门'],
            '项目经理': info['项目经理'],
            '项目成员': info['项目成员'],
            'TEKO链接': f"https://teko.woa.com/event/ai-agent/{agent_id}",
            '综合评分': '',
            '评委建议': '',
            '特点': '',
            '技术方案': '',
            '部门/个人': '',
            '类别': '',
            '特色简介': '',
            '业务价值': '',
            '有价值评论': '',
            '缺少字段': missing_info.get(agent_id, ''),
            '是否已评论': '否',
        }
        
        # 合并评分数据
        if agent_id in score_by_id:
            score_info = score_by_id[agent_id]
            merged[agent_id]['综合评分'] = score_info.get('综合评分', '')
            merged[agent_id]['评委建议'] = score_info.get('评委建议', '')
            merged[agent_id]['特点'] = score_info.get('特点', '')
            merged[agent_id]['技术方案'] = score_info.get('技术方案', '') or merged[agent_id]['技术方案']
        
        # 合并评论报告数据
        if agent_id in comment_data:
            comment_info = comment_data[agent_id]
            merged[agent_id]['部门/个人'] = comment_info.get('部门/个人', '')
            merged[agent_id]['类别'] = comment_info.get('类别', '')
            merged[agent_id]['特色简介'] = comment_info.get('特色简介', '')
            merged[agent_id]['业务价值'] = comment_info.get('业务价值', '')
            merged[agent_id]['有价值评论'] = comment_info.get('有价值评论', '')
            merged[agent_id]['是否已评论'] = '是'
            if comment_info.get('技术方案'):
                merged[agent_id]['技术方案'] = comment_info.get('技术方案', '')
            if comment_info.get('体验链接') and not merged[agent_id]['体验链接']:
                merged[agent_id]['体验链接'] = comment_info.get('体验链接', '')
    
    return merged

# ============ 6. 生成CSV报告 ============
def generate_csv(data):
    """生成CSV格式报告"""
    fields = ['id', 'agent名', 'TEKO链接', '体验链接', '主导部门', '涉及部门', 
              '项目经理', '项目成员', '综合评分', '评委建议', '特点', '技术方案',
              '部门/个人', '类别', '特色简介', '业务价值', '有价值评论', '缺少字段', '是否已评论']
    
    with open('AI_Agent大赛_部门赛_完整数据.csv', 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        writer.writeheader()
        for agent_id in sorted(data.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            writer.writerow(data[agent_id])
    
    print(f"CSV报告已生成: AI_Agent大赛_部门赛_完整数据.csv ({len(data)}条)")

# ============ 7. 生成Excel报告 ============
def generate_excel(data):
    """生成Excel多页签报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("openpyxl未安装，跳过Excel生成")
        return
    
    wb = openpyxl.Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== 页签1: 摘要 =====
    ws1 = wb.active
    ws1.title = "摘要"
    
    # 统计数据
    total = len(data)
    has_score = len([d for d in data.values() if d.get('综合评分')])
    has_link = len([d for d in data.values() if d.get('体验链接')])
    has_comment = len([d for d in data.values() if d.get('是否已评论') == '是'])
    missing_link = len([d for d in data.values() if '体验链接' in d.get('缺少字段', '')])
    missing_member = len([d for d in data.values() if '项目成员' in d.get('缺少字段', '')])
    
    # 计算平均分
    scores = [int(d['综合评分']) for d in data.values() if d.get('综合评分') and d['综合评分'].isdigit()]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    summary_data = [
        ['AI Agent大赛 - 部门赛道综合报告'],
        [f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'],
        [''],
        ['统计指标', '数值', '说明'],
        ['作品总数', total, '清单中的作品数量'],
        ['已评分项目', has_score, '有综合评分的项目'],
        ['平均评分', f'{avg_score:.1f}', '已评分项目的平均分'],
        ['有体验链接', has_link, f'占比 {has_link/total*100:.1f}%'],
        ['已评论作品', has_comment, f'占比 {has_comment/total*100:.1f}%'],
        ['缺少体验链接', missing_link, '需要补充'],
        ['缺少项目成员', missing_member, '需要补充'],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=16)
            elif row_idx == 4:
                cell.font = header_font
                cell.fill = header_fill
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 30
    
    # ===== 页签2: 已评分项目 =====
    ws2 = wb.create_sheet("已评分项目")
    scored_items = [(k, v) for k, v in data.items() if v.get('综合评分')]
    scored_items.sort(key=lambda x: int(x[1]['综合评分']) if x[1]['综合评分'].isdigit() else 0, reverse=True)
    
    headers = ['排名', 'ID', '作品名', '评分', '主导部门', '项目经理', '评委建议', '特点', '技术方案', '体验链接']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, (agent_id, info) in enumerate(scored_items, 2):
        ws2.cell(row=row_idx, column=1, value=row_idx-1)
        ws2.cell(row=row_idx, column=2, value=agent_id)
        ws2.cell(row=row_idx, column=3, value=info['agent名'])
        ws2.cell(row=row_idx, column=4, value=info['综合评分'])
        ws2.cell(row=row_idx, column=5, value=info['主导部门'])
        ws2.cell(row=row_idx, column=6, value=info['项目经理'])
        ws2.cell(row=row_idx, column=7, value=info['评委建议'])
        ws2.cell(row=row_idx, column=8, value=info['特点'])
        ws2.cell(row=row_idx, column=9, value=info['技术方案'])
        ws2.cell(row=row_idx, column=10, value=info['体验链接'])
    
    # 调整列宽
    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 8
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 20
    ws2.column_dimensions['I'].width = 30
    ws2.column_dimensions['J'].width = 40
    
    # ===== 页签3: 完整数据 =====
    ws3 = wb.create_sheet("完整数据")
    all_headers = ['ID', '作品名', 'TEKO链接', '体验链接', '主导部门', '涉及部门', 
                   '项目经理', '项目成员', '综合评分', '评委建议', '特点', '技术方案',
                   '类别', '特色简介', '业务价值', '有价值评论', '缺少字段', '是否已评论']
    
    for col, header in enumerate(all_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, agent_id in enumerate(sorted(data.keys(), key=lambda x: int(x) if x.isdigit() else 0), 2):
        info = data[agent_id]
        ws3.cell(row=row_idx, column=1, value=agent_id)
        ws3.cell(row=row_idx, column=2, value=info['agent名'])
        ws3.cell(row=row_idx, column=3, value=info['TEKO链接'])
        ws3.cell(row=row_idx, column=4, value=info['体验链接'])
        ws3.cell(row=row_idx, column=5, value=info['主导部门'])
        ws3.cell(row=row_idx, column=6, value=info['涉及部门'])
        ws3.cell(row=row_idx, column=7, value=info['项目经理'])
        ws3.cell(row=row_idx, column=8, value=info['项目成员'])
        ws3.cell(row=row_idx, column=9, value=info['综合评分'])
        ws3.cell(row=row_idx, column=10, value=info['评委建议'])
        ws3.cell(row=row_idx, column=11, value=info['特点'])
        ws3.cell(row=row_idx, column=12, value=info['技术方案'])
        ws3.cell(row=row_idx, column=13, value=info['类别'])
        ws3.cell(row=row_idx, column=14, value=info['特色简介'])
        ws3.cell(row=row_idx, column=15, value=info['业务价值'])
        ws3.cell(row=row_idx, column=16, value=info['有价值评论'])
        ws3.cell(row=row_idx, column=17, value=info['缺少字段'])
        ws3.cell(row=row_idx, column=18, value=info['是否已评论'])
    
    # ===== 页签4: 缺少信息 =====
    ws4 = wb.create_sheet("缺少信息")
    missing_items = [(k, v) for k, v in data.items() if v.get('缺少字段')]
    
    headers4 = ['ID', '作品名', '缺少字段', '主导部门', '项目经理', 'TEKO链接']
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    
    for row_idx, (agent_id, info) in enumerate(missing_items, 2):
        ws4.cell(row=row_idx, column=1, value=agent_id)
        ws4.cell(row=row_idx, column=2, value=info['agent名'])
        ws4.cell(row=row_idx, column=3, value=info['缺少字段'])
        ws4.cell(row=row_idx, column=4, value=info['主导部门'])
        ws4.cell(row=row_idx, column=5, value=info['项目经理'])
        ws4.cell(row=row_idx, column=6, value=info['TEKO链接'])
    
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 35
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 30
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 45
    
    # ===== 页签5: 已评论作品 =====
    ws5 = wb.create_sheet("已评论作品")
    commented_items = [(k, v) for k, v in data.items() if v.get('是否已评论') == '是']
    
    headers5 = ['ID', '作品名', '类别', '特色简介', '技术方案', '业务价值', '有价值评论', '体验链接']
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    for row_idx, (agent_id, info) in enumerate(commented_items, 2):
        ws5.cell(row=row_idx, column=1, value=agent_id)
        ws5.cell(row=row_idx, column=2, value=info['agent名'])
        ws5.cell(row=row_idx, column=3, value=info['类别'])
        ws5.cell(row=row_idx, column=4, value=info['特色简介'])
        ws5.cell(row=row_idx, column=5, value=info['技术方案'])
        ws5.cell(row=row_idx, column=6, value=info['业务价值'])
        ws5.cell(row=row_idx, column=7, value=info['有价值评论'])
        ws5.cell(row=row_idx, column=8, value=info['体验链接'])
    
    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 35
    ws5.column_dimensions['C'].width = 20
    ws5.column_dimensions['D'].width = 50
    ws5.column_dimensions['E'].width = 40
    ws5.column_dimensions['F'].width = 40
    ws5.column_dimensions['G'].width = 50
    ws5.column_dimensions['H'].width = 45
    
    # 保存
    wb.save('AI_Agent大赛_部门赛_综合报告.xlsx')
    print(f"Excel报告已生成: AI_Agent大赛_部门赛_综合报告.xlsx")

# ============ 8. 生成HTML报告 ============
def generate_html(data):
    """生成HTML网页报告"""
    # 统计数据
    total = len(data)
    has_score = len([d for d in data.values() if d.get('综合评分')])
    has_link = len([d for d in data.values() if d.get('体验链接')])
    has_comment = len([d for d in data.values() if d.get('是否已评论') == '是'])
    missing_link = len([d for d in data.values() if '体验链接' in d.get('缺少字段', '')])
    missing_member = len([d for d in data.values() if '项目成员' in d.get('缺少字段', '')])
    
    scores = [int(d['综合评分']) for d in data.values() if d.get('综合评分') and d['综合评分'].isdigit()]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    # 已评分项目
    scored_items = [(k, v) for k, v in data.items() if v.get('综合评分')]
    scored_items.sort(key=lambda x: int(x[1]['综合评分']) if x[1]['综合评分'].isdigit() else 0, reverse=True)
    
    # 已评论项目
    commented_items = [(k, v) for k, v in data.items() if v.get('是否已评论') == '是']
    
    # 缺少信息项目
    missing_items = [(k, v) for k, v in data.items() if v.get('缺少字段')]
    
    # 部门统计
    dept_stats = {}
    for info in data.values():
        dept = info.get('主导部门', '未知')
        if dept:
            bg = dept.split('-')[0] if '-' in dept else dept
            dept_stats[bg] = dept_stats.get(bg, 0) + 1
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AI Agent大赛 - 部门赛道综合报告</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background: #f5f7fa; color: #333; line-height: 1.6; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 40px 20px; text-align: center; margin-bottom: 30px; border-radius: 12px; }}
        header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        header p {{ opacity: 0.9; font-size: 1.1em; }}
        .tabs {{ display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 20px; background: white; padding: 15px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .tab {{ padding: 12px 24px; cursor: pointer; border: none; background: #e9ecef; border-radius: 8px; font-size: 1em; transition: all 0.3s; }}
        .tab:hover {{ background: #dee2e6; }}
        .tab.active {{ background: #667eea; color: white; }}
        .tab-content {{ display: none; background: white; padding: 30px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .tab-content.active {{ display: block; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .stat-card {{ background: white; padding: 25px; border-radius: 12px; text-align: center; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .stat-card h3 {{ font-size: 2.5em; color: #667eea; margin-bottom: 5px; }}
        .stat-card p {{ color: #666; font-size: 0.95em; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid #e9ecef; }}
        th {{ background: #667eea; color: white; font-weight: 600; position: sticky; top: 0; }}
        tr:hover {{ background: #f8f9fa; }}
        .score {{ display: inline-block; padding: 4px 12px; border-radius: 20px; font-weight: bold; }}
        .score-high {{ background: #d4edda; color: #155724; }}
        .score-mid {{ background: #fff3cd; color: #856404; }}
        .score-low {{ background: #f8d7da; color: #721c24; }}
        .link {{ color: #667eea; text-decoration: none; }}
        .link:hover {{ text-decoration: underline; }}
        .search-box {{ width: 100%; padding: 12px 20px; border: 2px solid #e9ecef; border-radius: 8px; font-size: 1em; margin-bottom: 20px; }}
        .search-box:focus {{ outline: none; border-color: #667eea; }}
        .tag {{ display: inline-block; padding: 2px 8px; background: #e9ecef; border-radius: 4px; font-size: 0.85em; margin: 2px; }}
        .missing {{ color: #dc3545; }}
        .analysis {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-top: 20px; }}
        .analysis h3 {{ color: #667eea; margin-bottom: 15px; }}
        .chart-container {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin-top: 20px; }}
        .chart {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .bar {{ height: 25px; background: linear-gradient(90deg, #667eea, #764ba2); border-radius: 4px; margin: 5px 0; transition: width 0.5s; }}
        .bar-label {{ display: flex; justify-content: space-between; font-size: 0.9em; color: #666; }}
        details {{ margin: 10px 0; }}
        summary {{ cursor: pointer; padding: 10px; background: #f8f9fa; border-radius: 8px; }}
        summary:hover {{ background: #e9ecef; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>🏆 AI Agent大赛 - 部门赛道综合报告</h1>
            <p>生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")} | 数据来源: 部门赛清单 + TEKO平台 + 体验评分表</p>
        </header>
        
        <div class="tabs">
            <button class="tab active" onclick="showTab('summary')">📊 摘要分析</button>
            <button class="tab" onclick="showTab('scored')">⭐ 已评分项目 ({has_score})</button>
            <button class="tab" onclick="showTab('commented')">💬 已评论作品 ({len(commented_items)})</button>
            <button class="tab" onclick="showTab('all')">📋 完整数据 ({total})</button>
            <button class="tab" onclick="showTab('missing')">⚠️ 缺少信息 ({len(missing_items)})</button>
            <button class="tab" onclick="showTab('dept')">🏢 部门分布</button>
        </div>
        
        <!-- 摘要分析 -->
        <div id="summary" class="tab-content active">
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>{total}</h3>
                    <p>作品总数</p>
                </div>
                <div class="stat-card">
                    <h3>{has_score}</h3>
                    <p>已评分项目</p>
                </div>
                <div class="stat-card">
                    <h3>{avg_score:.1f}</h3>
                    <p>平均评分</p>
                </div>
                <div class="stat-card">
                    <h3>{has_link}</h3>
                    <p>有体验链接</p>
                </div>
                <div class="stat-card">
                    <h3>{has_comment}</h3>
                    <p>已评论作品</p>
                </div>
                <div class="stat-card">
                    <h3>{missing_link}</h3>
                    <p>缺少体验链接</p>
                </div>
            </div>
            
            <div class="chart-container">
                <div class="chart">
                    <h4>📈 评分分布</h4>
                    {''.join([f'<div class="bar-label"><span>{k}分</span><span>{v}个</span></div><div class="bar" style="width: {v/max(1, max([len([s for s in scores if s >= i*10 and s < (i+1)*10]) for i in range(10)]))*100}%"></div>' for k, v in sorted({f'{i*10}-{(i+1)*10-1}': len([s for s in scores if s >= i*10 and s < (i+1)*10]) for i in range(10)}.items(), reverse=True) if v > 0])}
                </div>
                <div class="chart">
                    <h4>🏢 BG分布 Top 10</h4>
                    {''.join([f'<div class="bar-label"><span>{k}</span><span>{v}个</span></div><div class="bar" style="width: {v/max(dept_stats.values())*100}%"></div>' for k, v in sorted(dept_stats.items(), key=lambda x: x[1], reverse=True)[:10]])}
                </div>
            </div>
            
            <div class="analysis">
                <h3>📝 整体分析</h3>
                <ul>
                    <li><strong>数据完整性</strong>: {has_link/total*100:.1f}% 的项目有体验链接，{(total-missing_member)/total*100:.1f}% 有完整的项目成员信息</li>
                    <li><strong>评分情况</strong>: 已评分 {has_score} 个项目，平均分 {avg_score:.1f} 分，最高分 {max(scores) if scores else 0} 分</li>
                    <li><strong>评论覆盖</strong>: 已评论 {has_comment} 个作品，覆盖率 {has_comment/total*100:.1f}%</li>
                    <li><strong>部门分布</strong>: 涉及 {len(dept_stats)} 个BG，{max(dept_stats.items(), key=lambda x: x[1])[0] if dept_stats else '未知'} 贡献最多 ({max(dept_stats.values()) if dept_stats else 0} 个)</li>
                </ul>
            </div>
        </div>
        
        <!-- 已评分项目 -->
        <div id="scored" class="tab-content">
            <h2>⭐ 已评分项目 (按分数排序)</h2>
            <table>
                <thead>
                    <tr>
                        <th>排名</th>
                        <th>ID</th>
                        <th>作品名</th>
                        <th>评分</th>
                        <th>主导部门</th>
                        <th>评委建议</th>
                        <th>特点/技术</th>
                        <th>操作</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f'''<tr>
                        <td>{idx}</td>
                        <td>{agent_id}</td>
                        <td><strong>{info["agent名"]}</strong></td>
                        <td><span class="score {'score-high' if int(info['综合评分']) >= 90 else 'score-mid' if int(info['综合评分']) >= 70 else 'score-low'}">{info["综合评分"]}</span></td>
                        <td>{info["主导部门"][:20]}...</td>
                        <td>{info["评委建议"][:100]}...</td>
                        <td>{info.get("特点", "") or info.get("技术方案", "")[:50]}...</td>
                        <td><a class="link" href="{info["TEKO链接"]}" target="_blank">查看</a></td>
                    </tr>''' for idx, (agent_id, info) in enumerate(scored_items, 1)])}
                </tbody>
            </table>
        </div>
        
        <!-- 已评论作品 -->
        <div id="commented" class="tab-content">
            <h2>💬 已评论作品</h2>
            <input type="text" class="search-box" placeholder="搜索作品名、类别..." onkeyup="searchTable(this, 'commented-table')">
            <table id="commented-table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>作品名</th>
                        <th>类别</th>
                        <th>特色简介</th>
                        <th>技术方案</th>
                        <th>有价值评论</th>
                        <th>操作</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f'''<tr>
                        <td>{agent_id}</td>
                        <td><strong>{info["agent名"]}</strong></td>
                        <td>{''.join([f'<span class="tag">{c}</span>' for c in info.get("类别", "").split("、") if c])}</td>
                        <td>{info.get("特色简介", "")[:80]}...</td>
                        <td>{info.get("技术方案", "")[:60]}...</td>
                        <td>{info.get("有价值评论", "")[:80]}...</td>
                        <td><a class="link" href="{info["TEKO链接"]}" target="_blank">查看</a></td>
                    </tr>''' for agent_id, info in commented_items])}
                </tbody>
            </table>
        </div>
        
        <!-- 完整数据 -->
        <div id="all" class="tab-content">
            <h2>📋 完整数据</h2>
            <input type="text" class="search-box" placeholder="搜索作品名、部门、项目经理..." onkeyup="searchTable(this, 'all-table')">
            <div style="overflow-x: auto;">
                <table id="all-table">
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>作品名</th>
                            <th>主导部门</th>
                            <th>项目经理</th>
                            <th>评分</th>
                            <th>体验链接</th>
                            <th>状态</th>
                            <th>操作</th>
                        </tr>
                    </thead>
                    <tbody>
                        {''.join([f'''<tr>
                            <td>{agent_id}</td>
                            <td><strong>{info["agent名"]}</strong></td>
                            <td>{info["主导部门"][:25]}...</td>
                            <td>{info["项目经理"]}</td>
                            <td>{info["综合评分"] or "-"}</td>
                            <td>{"✅" if info["体验链接"] else "❌"}</td>
                            <td>{"<span class='missing'>缺: " + info["缺少字段"] + "</span>" if info["缺少字段"] else "✅"}</td>
                            <td><a class="link" href="{info["TEKO链接"]}" target="_blank">查看</a></td>
                        </tr>''' for agent_id, info in sorted(data.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0)])}
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- 缺少信息 -->
        <div id="missing" class="tab-content">
            <h2>⚠️ 缺少信息的项目</h2>
            <p style="color: #666; margin-bottom: 20px;">以下项目缺少体验链接或项目成员信息，需要补充完善。</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>作品名</th>
                        <th>缺少字段</th>
                        <th>主导部门</th>
                        <th>项目经理</th>
                        <th>操作</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f'''<tr>
                        <td>{agent_id}</td>
                        <td><strong>{info["agent名"]}</strong></td>
                        <td><span class="missing">{info["缺少字段"]}</span></td>
                        <td>{info["主导部门"][:25]}...</td>
                        <td>{info["项目经理"]}</td>
                        <td><a class="link" href="{info["TEKO链接"]}" target="_blank">查看</a></td>
                    </tr>''' for agent_id, info in missing_items])}
                </tbody>
            </table>
        </div>
        
        <!-- 部门分布 -->
        <div id="dept" class="tab-content">
            <h2>🏢 部门分布统计</h2>
            <div class="chart-container">
                <div class="chart" style="grid-column: span 2;">
                    <h4>各BG作品数量</h4>
                    {''.join([f'<div class="bar-label"><span>{k}</span><span>{v}个 ({v/total*100:.1f}%)</span></div><div class="bar" style="width: {v/max(dept_stats.values())*100}%"></div>' for k, v in sorted(dept_stats.items(), key=lambda x: x[1], reverse=True)])}
                </div>
            </div>
        </div>
    </div>
    
    <script>
        function showTab(tabId) {{
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            event.target.classList.add('active');
        }}
        
        function searchTable(input, tableId) {{
            const filter = input.value.toLowerCase();
            const table = document.getElementById(tableId);
            const rows = table.getElementsByTagName('tr');
            for (let i = 1; i < rows.length; i++) {{
                const text = rows[i].textContent.toLowerCase();
                rows[i].style.display = text.includes(filter) ? '' : 'none';
            }}
        }}
    </script>
</body>
</html>'''
    
    with open('AI_Agent大赛_部门赛_综合报告.html', 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"HTML报告已生成: AI_Agent大赛_部门赛_综合报告.html")

# ============ 主函数 ============
def main():
    print("=" * 50)
    print("AI Agent大赛综合报告生成器")
    print("=" * 50)
    
    # 整合数据
    print("\n[1/4] 整合所有数据源...")
    data = merge_all_data()
    print(f"    - 共整合 {len(data)} 个部门赛作品")
    
    # 生成CSV
    print("\n[2/4] 生成CSV报告...")
    generate_csv(data)
    
    # 生成Excel
    print("\n[3/4] 生成Excel报告...")
    generate_excel(data)
    
    # 生成HTML
    print("\n[4/4] 生成HTML报告...")
    generate_html(data)
    
    print("\n" + "=" * 50)
    print("报告生成完成！")
    print("=" * 50)

if __name__ == '__main__':
    main()
